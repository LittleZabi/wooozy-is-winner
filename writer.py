import datetime
from setting import resultsFolder, base_dir
from openpyxl import Workbook, load_workbook,utils,worksheet
from _lib import log
import os


class typeWriter:
    def __init__(self) -> None:
        self.file_cols = ['card name', 'time', 'card url text', 'number of runners', 'verdict', 'tv',
                          'date', 'forecast', 'surface', 'verdict players', 'horses list']
        self.filename = self.fname()
        self.horse_cols = ['runner name', 'runner form', 'last run','cdbf', 'tips']
        self.ExcelSheet()
        self.fileCounter = 1

    def ExcelSheet(self):
        log('Creating excel spreadsheet...', 'alert')
        wb = Workbook()
        ws = False
        for sheet in wb:
            if sheet.title == 'list':
                ws = sheet
                break
        if ws == False:
            ws = wb.create_sheet('list')
        ws.cell(row=1, column=1, value='#')
        ws.cell(row=1, column=2, value='filename')
        ws.cell(row=1, column=3, value='link')
        wb.save(self.filename)

    def onExcel(self, data):
        wb = load_workbook(self.filename)
        ws = False
        listSheet = wb['list']
        for sheet in wb:
            if sheet.title == data['card name']:
                ws = sheet
                break
        if ws == False:
            ws = wb.create_sheet(data['card name'])
        log('Saving in excel sheet....')
        fileName = self.filename.split('\\')
        fileName = fileName[len(fileName) -1]
        fileName = fileName.replace('\\', '').replace('/', '')
        link = f"{fileName}#'{ws.title}'!A1"
        listSheetMaxRow = listSheet.max_row+1
        listSheet.cell(row=listSheetMaxRow, column=1, value=listSheetMaxRow)
        listSheet.cell(row=listSheetMaxRow, column=2, value=data['card name'])
        listSheet.cell(row=listSheetMaxRow, column=3).hyperlink = link
        listSheet.cell(row=listSheetMaxRow, column=3).value = f"{ws.title}"
        listSheet.cell(row=listSheetMaxRow, column=3).style = "Hyperlink"
        i = 1
        for cols in self.file_cols:
            ws.cell(row=1, column=i, value=cols)
            i += 1
        i = 1
        __row__ = ws.max_row + 1
        klash = len(data.keys()) + 1
        ws.cell(row=1, column=klash).hyperlink = f"{fileName}#'list'!A1"
        ws.cell(row=1, column=klash).value = f"List of files"
        ws.cell(row=1, column=klash).style = "Hyperlink"
        for row in data.keys():
            if row != 'horses_list':
                ws.cell(row=__row__, column=i, value=data[row])
            if row == 'horses_list':
                log('Horses detail saving...')
                ews = f"Horses-{data['card name']}-{data['time'].strip().replace(':', '-')}".replace(' ', '-').replace('(', '_').replace(')', '_')
                horsesFileName = ews
                colx = utils.cell.get_column_letter(i)
                link = f"{fileName}#'{ews}'!A1"
                ws.cell(row=__row__, column=i).hyperlink = link
                ws.cell(row=__row__, column=i).value = f"Refer {ews}"
                ws.cell(row=__row__, column=i).style = "Hyperlink"
                ews = wb.create_sheet(ews)
                k = 1
                for cols in self.horse_cols:
                    ews.cell(row=1, column=k, value=cols)
                    k += 1
                __erow__ = 1
                for erow in data['horses_list']:
                    k = 1
                    __erow__ = __erow__ + 1
                    for hrs_row in erow.keys():
                        ews.cell(row=__erow__, column=k, value=erow[hrs_row])
                        k += 1
                ref_row = len(erow.keys()) + 1
                link = f"{fileName}#'{data['card name']}'!{colx}{__row__}"
                ews.cell(row=1, column=ref_row).hyperlink = link
                ews.cell(row=1, column=ref_row).value = f"Visit {ws.title}"
                ews.cell(row=1, column=ref_row).style = "Hyperlink"
                ews.cell(row=1, column=ref_row+1).hyperlink = f"{fileName}#'list'!A1"
                ews.cell(row=1, column=ref_row+1).value = "List of files"
                ews.cell(row=1, column=ref_row+1).style = "Hyperlink"
                
                listSheetMaxRow = listSheet.max_row+1
                listSheet.cell(row=listSheetMaxRow, column=1, value=listSheetMaxRow)
                listSheet.cell(row=listSheetMaxRow, column=2, value=horsesFileName)
                listSheet.cell(row=listSheetMaxRow, column=3).hyperlink = link
                listSheet.cell(row=listSheetMaxRow, column=3).value = f"{ews.title}"
                listSheet.cell(row=listSheetMaxRow, column=3).style = "Hyperlink"
                
                log('Horses detail saved successfully.')
            i += 1
        wb.save(self.filename)
        log('Saved in excel file ' + self.filename)

    def fname(self):
        now = datetime.datetime.now()
        ext = '.xlsx'
        return f"{base_dir}/{resultsFolder}/{now.strftime('%d-%m-%Y_%H-%M-%S')}{ext}"

if __name__ == '__main__':
    pass
    # write = typeWriter()
    # data = {'card name': 'Lingfield (AW)', 'time': '1:20', 'card url text': 'talkSPORT Powered By Fans Handicap', 'number of runners': '8 runners', 'date': '24 Feb 2023', 'tv': 'SKY', 'surface': 'Polytrack', 'forecast': 'Betting Forecast 2/1 All Dunn, 4/1 Regal Glory, 5/1 Cariad, 7/1 Man Made Of Smoke, 8/1 Stintino Sunset, Twilight Kiss, 12/1 Deed Pole, 20/1 Spice Rack.', 'runner form': '7121-4', 'last run': '21', 'cdbf': '', 'tips': '1 tip',
    #         'verdict': 'ALL DUNN has found some improvement on AW in recent weeks, making it 3 wins from his last 4 starts in a C&D handicap 2 weeks ago, and a 4 lb higher mark ought not prevent another bold showing on that evidence. Regal Glory is also going through a good spell and is worth a look upped to 1m. Stintino Sunset and Cariad complete the shortlist.', 'verdict players': 'All Dunn => 11/8 | Regal Glory => 9/2 | Stintino Sunset => 16/1 | ',
    #         'horses_list': [{'runner name': 'Nickelsonthedime', 'runner form': '1353P3', 'last run': '19', 'cdbf': 'C', 'tips': '3 tips'}, {'runner name': 'Cafe Pushkin', 'runner form': '285-67', 'last run': '85', 'cdbf': '', 'tips': '2 tips'}, {'runner name': 'Dame Prestige', 'runner form': '5-35P5', 'last run': '11', 'cdbf': '', 'tips': ''}, {'runner name': 'Design Icon', 'runner form': '1-6902', 'last run': '16', 'cdbf': '', 'tips': '3 tips'}, {'runner name': 'Turnaway', 'runner form': '70-410', 'last run': '24', 'cdbf': '', 'tips': ''}, {'runner name': "Ailes D'Amour", 'runner form': '57P-96', 'last run': '22', 'cdbf': '', 'tips': ''}, {'runner name': 'Seymour Promise', 'runner form': '9P0-22', 'last run': '22', 'cdbf': '', 'tips': '5 tips'}, {'runner name': 'On The Bandwagon', 'runner form': '5P6427', 'last run': '23', 'cdbf': '', 'tips': ''}, {'runner name': 'Mr Palmtree', 'runner form': '1443/5', 'last run': '17', 'cdbf': 'D', 'tips': ''}, {'runner name': 'Lord Bill', 'runner form': '8-P447', 'last run': '16', 'cdbf': '', 'tips': ''}, {'runner name': 'Telefina', 'runner form': '9P2530', 'last run': '29', 'cdbf': '', 'tips': ''}]
    #         }
    # write.onExcel(data)